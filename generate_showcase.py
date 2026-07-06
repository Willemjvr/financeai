"""Generate showcase documents for FINANCEAI demo to a CA.

Creates:
  - showcase/client_list.xlsx   — client records with PII
  - showcase/invoice.pdf         — invoice with client details
  - showcase/bank_statement.pdf  — bank statement with account info
All documents in: ~/repos/financeai/showcase/
"""

import os, textwrap
from pathlib import Path

OUT_DIR = Path.home() / "repos" / "financeai" / "showcase"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ──────────────────────────────────────────
# Fake data (keep it latin-1 safe for PDF)
# ──────────────────────────────────────────

clients = [
    {"name": "Thabo Mokoena",       "id_num": "880101 1234 081", "phone": "081 487 1584", "email": "thabo.mokoena@gmail.com",       "balance": 45230.00, "type": "Business"},
    {"name": "Jane Doe",            "id_num": "900215 5678 092", "phone": "082 555 1234", "email": "jane.doe@example.com",          "balance": 12800.50, "type": "Personal"},
    {"name": "Sarah van der Merwe", "id_num": "850712 3456 073", "phone": "072 987 6543", "email": "sarah.vdm@webmail.co.za",      "balance": 78500.75, "type": "Business"},
    {"name": "Blessing Nkosi",      "id_num": "920930 7890 064", "phone": "083 123 4567", "email": "bnkosi@outlook.com",           "balance": 5200.00,  "type": "Personal"},
    {"name": "Ahmed Patel",         "id_num": "780504 2345 085", "phone": "071 555 9876", "email": "ahmed.patel@consulting.co.za", "balance": 120300.00, "type": "Business"},
]

invoices = [
    ("Monthly bookkeeping - March 2024",           1, 4500.00),
    ("VAT return preparation - Feb 2024",          1, 1200.00),
    ("Payroll processing (12 employees)",           1, 2800.00),
    ("Tax consulting - corporate restructuring",   2, 1750.00),
    ("Expense reconciliation and coding",          4, 650.00),
]

transactions = [
    ("2024-03-01", "Salary deposit - Overgrowth AI",              45000.00,  "CR"),
    ("2024-03-03", "EFT - Makro Office Supplies",                  -2340.50,  "DR"),
    ("2024-03-07", "Debit order - FNB Business Insurance",         -1250.00,  "DR"),
    ("2024-03-10", "EFT - Client payment INV-2024-0039 (J Doe)",   12800.50,  "CR"),
    ("2024-03-14", "Card purchase - Takealot Business",             -899.99,  "DR"),
    ("2024-03-18", "EFT - SARS VAT payment",                      -11250.00,  "DR"),
    ("2024-03-22", "Salary deposit - Overgrowth AI",              45000.00,  "CR"),
    ("2024-03-25", "EFT - Client payment INV-2024-0040 (A Patel)", 15200.00,  "CR"),
    ("2024-03-28", "Debit order - Microsoft 365 Business",          -450.00,  "DR"),
    ("2024-03-30", "EFT - Rent payment (The Business Hub)",      -12000.00,  "DR"),
]

# ──────────────────────────────────────────
# 1. Client List - Excel
# ──────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Client List"

ws.merge_cells("A1:F1")
ws["A1"] = "OVERGROWTH AI - Client List"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center")

headers = ["Client Name", "SA ID Number", "Phone", "Email", "Account Balance", "Account Type"]
hf = PatternFill("solid", fgColor="D6E4F0")
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, size=11)
    c.fill = hf
    c.alignment = Alignment(horizontal="center")

for i, cl in enumerate(clients, 3):
    ws.cell(row=i, column=1, value=cl["name"])
    ws.cell(row=i, column=2, value=cl["id_num"])
    ws.cell(row=i, column=3, value=cl["phone"])
    ws.cell(row=i, column=4, value=cl["email"])
    c = ws.cell(row=i, column=5, value=cl["balance"])
    c.number_format = '#,##0.00'
    ws.cell(row=i, column=6, value=cl["type"])

ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 14

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=2, max_row=len(clients)+2, min_col=1, max_col=6):
    for cell in row:
        cell.border = border

xlsx_path = OUT_DIR / "client_list.xlsx"
wb.save(str(xlsx_path))
print(f"OK Created {xlsx_path.name}")

# ──────────────────────────────────────────
# 2. Invoice - PDF
# ──────────────────────────────────────────

from fpdf import FPDF

class InvoicePDF(FPDF):
    def header(self):
        self.set_fill_color(31, 78, 121)
        self.rect(0, 0, 210, 32, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 20)
        self.set_y(6)
        self.cell(0, 10, "OVERGROWTH AI", align="C")
        self.set_font("Helvetica", "", 9)
        self.set_y(17)
        self.cell(0, 6, "Financial Services  |  Tax & Accounting  |  Business Consulting", align="C")
        self.set_y(23)
        self.cell(0, 6, "15 St Georges Mall, Cape Town, 8001  |  VAT: 4870234567", align="C")
        self.ln(36)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

cl = clients[0]
pdf = InvoicePDF()
pdf.alias_nb_pages()
pdf.add_page()

# Invoice title
pdf.set_text_color(31, 78, 121)
pdf.set_font("Helvetica", "B", 15)
pdf.cell(0, 10, "INVOICE")
pdf.ln(14)

# Details
pdf.set_text_color(0, 0, 0)
pdf.set_font("Helvetica", "", 10)
details = [
    ("Invoice #:", "INV-2024-0042"),
    ("Date:", "25 March 2024"),
    ("Due Date:", "15 April 2024"),
]
for label, val in details:
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(40, 6, label)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, val)
    pdf.ln()

pdf.ln(6)

# Bill to
pdf.set_font("Helvetica", "B", 10)
pdf.cell(0, 6, "Bill To:")
pdf.ln()
pdf.set_font("Helvetica", "", 10)
for line in [cl["name"], "12 High Street, Sandton, 2196", f"Email: {cl['email']}", f"Phone: {cl['phone']}", f"SA ID: {cl['id_num']}"]:
    pdf.cell(0, 6, line)
    pdf.ln()
pdf.ln(6)

# Table
col_w = [78, 18, 28, 28, 28]
pdf.set_fill_color(31, 78, 121)
pdf.set_text_color(255, 255, 255)
pdf.set_font("Helvetica", "B", 9)
for i, h in enumerate(["Description", "Qty", "Unit (R)", "VAT (R)", "Total (R)"]):
    pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
pdf.ln()

pdf.set_text_color(0, 0, 0)
pdf.set_font("Helvetica", "", 9)
subtotal = 0
for item, qty, rate in invoices:
    line_total = qty * rate
    vat = line_total * 0.15
    subtotal += line_total
    row = [item, str(qty), f"{rate:,.2f}", f"{vat:,.2f}", f"{line_total:,.2f}"]
    for i, val in enumerate(row):
        pdf.cell(col_w[i], 7, val, border=1, align="L" if i == 0 else "C")
    pdf.ln()

# Totals
vat_total = subtotal * 0.15
grand = subtotal + vat_total
pdf.ln(3)
pdf.set_font("Helvetica", "B", 10)
pdf.cell(160, 7, "Subtotal:", align="R")
pdf.cell(0, 7, f"R {subtotal:,.2f}", align="R")
pdf.ln()
pdf.cell(160, 7, "VAT (15%):", align="R")
pdf.cell(0, 7, f"R {vat_total:,.2f}", align="R")
pdf.ln()
pdf.set_fill_color(31, 78, 121)
pdf.set_text_color(255, 255, 255)
pdf.cell(160, 8, "TOTAL DUE:", align="R", fill=True)
pdf.cell(0, 8, f"R {grand:,.2f}", align="R", fill=True)
pdf.ln(15)

# Payment details
pdf.set_text_color(100, 100, 100)
pdf.set_font("Helvetica", "I", 8)
pdf.cell(0, 5, "Payment: FNB Business Account 6270 1234 5678  |  Branch: 255005", align="C")
pdf.ln()
pdf.cell(0, 5, "Thank you for your business.", align="C")

pdf_path = OUT_DIR / "invoice.pdf"
pdf.output(str(pdf_path))
print(f"OK Created {pdf_path.name}")

# ──────────────────────────────────────────
# 3. Bank Statement - PDF
# ──────────────────────────────────────────

class StmtPDF(FPDF):
    def header(self):
        self.set_fill_color(0, 104, 56)
        self.rect(0, 0, 210, 28, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 18)
        self.set_y(6)
        self.cell(0, 10, "FNB BUSINESS ACCOUNT STATEMENT", align="C")
        self.set_font("Helvetica", "", 8)
        self.set_y(17)
        self.cell(0, 6, "First National Bank  |  www.fnb.co.za", align="C")
        self.ln(33)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, "Demo document - not a real bank statement", align="C")

pdf2 = StmtPDF()
pdf2.alias_nb_pages()
pdf2.add_page()

# Account info
pdf2.set_text_color(0, 0, 0)
info_lines = [
    ("Account Holder:", cl["name"]),
    ("Account Number:", "6270 1234 5678"),
    ("Branch Code:", "255005"),
    ("Statement Period:", "01 March 2024 - 31 March 2024"),
    ("SA ID Number:", cl["id_num"]),
]
for label, val in info_lines:
    pdf2.set_font("Helvetica", "B", 10)
    pdf2.cell(45, 6, label)
    pdf2.set_font("Helvetica", "", 10)
    pdf2.cell(0, 6, val)
    pdf2.ln()
pdf2.ln(8)

# Transactions table
sc = [30, 82, 38, 22]
pdf2.set_fill_color(0, 104, 56)
pdf2.set_text_color(255, 255, 255)
pdf2.set_font("Helvetica", "B", 9)
for i, h in enumerate(["Date", "Description", "Amount (R)", "Type"]):
    pdf2.cell(sc[i], 7, h, border=1, align="C", fill=True)
pdf2.ln()

pdf2.set_text_color(0, 0, 0)
pdf2.set_font("Helvetica", "", 9)
bal = 15000.00
for idx, (date, desc, amt, typ) in enumerate(transactions):
    fill = idx % 2 == 0
    bal += amt
    pdf2.cell(sc[0], 6, date, border=1, align="C", fill=fill)
    pdf2.cell(sc[1], 6, desc, border=1, fill=fill)
    pdf2.cell(sc[2], 6, f"{amt:+,.2f}", border=1, align="R", fill=fill)
    pdf2.cell(sc[3], 6, typ, border=1, align="C", fill=fill)
    pdf2.ln()

# Closing
pdf2.set_font("Helvetica", "B", 10)
pdf2.cell(sc[0]+sc[1], 8, "Closing Balance:  ", border=1, align="R")
pdf2.cell(sc[2], 8, f"R {bal:,.2f}", border=1, align="C", fill=True)
pdf2.cell(sc[3], 8, "", border=1)

stmt_path = OUT_DIR / "bank_statement.pdf"
pdf2.output(str(stmt_path))
print(f"OK Created {stmt_path.name}")

print(f"\nAll documents in: {OUT_DIR}")
